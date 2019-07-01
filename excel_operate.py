import pandas as pd
from pandas import DataFrame
import os
import sys
from openpyxl import load_workbook


def get_visible_names(file):
    wb = load_workbook(file)
    visible_sheets = [sheet for sheet in wb.sheetnames if wb[sheet].sheet_state == 'visible']
    return visible_sheets


# 获取保留的属性
def get_keys(file, sheet_name):
    # out_file = os.path.join('excels','out.xlsx')
    # out_sheet_name = '明细汇总'
    # out_sheet_name = get_visible_names(out_file)
    out_frame = pd.read_excel(file, sheet_name=sheet_name)
    keys1 = out_frame.keys().values
    keys_out = list(keys1.tolist())
    # 过滤掉没有名字的
    keys_out = [key for key in keys_out if not 'Unnamed:' in key]
    print('get keys {}'.format(keys_out))
    return keys_out


def write_file_by_keys(files, store_keys):
    out_frame = DataFrame(columns=store_keys)
    for file in files:
        sheet_name = get_visible_names(file)
        if isinstance(sheet_name, list):
            for sheet in sheet_name:
                f = pd.read_excel(file, sheet_name=sheet)
                keys_have = f.keys().values.tolist()
                keys_use = [key for key in store_keys if key in keys_have]
                tmp = f[keys_use]
                out_frame = pd.concat([out_frame, tmp], sort=False)
    output_excel = os.path.join('static', 'output.xlsx')
    out_frame.to_excel(output_excel,columns=store_keys)
    print('write file {} with keys {}'.format(output_excel,keys_use))

def test_print_excel_keys(excel):
    for s in get_visible_names(excel):
        get_keys(excel,s)
    
if __name__ == '__main__':
    test_print_excel_keys(sys.argv[1])
